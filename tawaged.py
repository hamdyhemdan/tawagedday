import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, time
import re
from openpyxl.styles import PatternFill

st.set_page_config(page_title="نظام الدمج السريع الشامل", layout="centered")

def parse_manual_time(time_str):
    try:
        return datetime.strptime(time_str.strip(), "%H:%M").time()
    except:
        return None

def smart_read_excel(uploaded_file):
    try:
        if uploaded_file.name.endswith('.xls'):
            return pd.read_excel(uploaded_file, engine='xlrd')
        else:
            return pd.read_excel(uploaded_file)
    except Exception as e:
        try:
            return pd.read_excel(uploaded_file)
        except:
            raise ValueError(f"لم نتمكن من قراءة الملف {uploaded_file.name}. الخطأ المباشر: {e}")

def main():
    st.title("🚀 نظام دمج وتلوين البصمة الشامل")
    st.write("النسخة المحدثة: تدعم كافة صيغ الإكسيل (xls / xlsx) مع إخفاء أعمدة (اساسى والشيفت) تلقائياً")

    main_file = st.file_uploader("1. ملف البصمة الأساسي (AC-LOG)", type=["xlsx", "xls"])
    source_file = st.file_uploader("2. ملف بيانات المصدر (الجدول)", type=["xlsx", "xls"])

    st.divider()
    tc1, tc2 = st.columns(2)
    with tc1:
        start_t_raw = st.text_input("من وقت (مثلاً 08:00):", value="08:00")
    with tc2:
        end_t_raw = st.text_input("إلى وقت (مثلاً 08:15):", value="08:15")

    if st.button("🚀 ابدأ المعالجة الآن"):
        if main_file and source_file:
            try:
                with st.spinner('جاري قراءة البيانات ودمج الأعمدة وتطبيق التلوين...'):
                    df_main = smart_read_excel(main_file)
                    df_source = smart_read_excel(source_file)
                    
                    df_main.columns = df_main.columns.str.strip()
                    df_source.columns = df_source.columns.str.strip()
                    
                    # أسماء الأعمدة للربط
                    col_id_1, col_date_1 = "رقم البصمه", "التاريخ"
                    col_id_2, col_date_2 = "Badgenumber", "Date"
                    target_color_col = "الوقت"
                    
                    if col_id_1 not in df_main.columns:
                        st.error(f"❌ خطأ: لم يتم العثور على عمود '{col_id_1}' في الملف الأول.")
                        return
                    if col_id_2 not in df_source.columns:
                        st.error(f"❌ خطأ: لم يتم العثور على عمود '{col_id_2}' في ملف الجدول الثاني.")
                        return
                    
                    clean_id = lambda v: str(int(float(v))).strip() if pd.notna(v) else ""
                    def extract_day(v):
                        if isinstance(v, (datetime, pd.Timestamp)): return str(v.day)
                        m = re.search(r'(\d+)', str(v))
                        return str(int(m.group(1))) if m else "NONE"

                    df_main['M_ID'] = df_main[col_id_1].apply(clean_id)
                    df_main['M_DAY'] = df_main[col_date_1].apply(extract_day)
                    df_source['M_ID'] = df_source[col_id_2].apply(clean_id)
                    df_source['M_DAY'] = df_source[col_date_2].apply(extract_day)

                    # دمج البيانات (بشكل تلقائي سيتم تضمين Name, Department, SHIFT لأنهم في df_source)
                    cols_to_bring = [c for c in df_source.columns if c not in ['M_ID', 'M_DAY', col_id_2, col_date_2]]
                    df_src_subset = df_source[['M_ID', 'M_DAY'] + cols_to_bring].drop_duplicates(subset=['M_ID', 'M_DAY'])
                    result_df = pd.merge(df_main, df_src_subset, on=['M_ID', 'M_DAY'], how='left')
                    result_df = result_df.drop(columns=['M_ID', 'M_DAY'])

                    # حذف الأعمدة غير المرغوب فيها (مع إضافة SHIFT للحذف)
                    columns_to_drop = ["اساسى", "الشيفت", "SHIFT"]
                    for col in columns_to_drop:
                        if col in result_df.columns:
                            result_df = result_df.drop(columns=[col])

                    start_t = parse_manual_time(start_t_raw)
                    end_t = parse_manual_time(end_t_raw)
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                        worksheet = writer.sheets['Sheet1']
                        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        
                        if target_color_col in result_df.columns:
                            col_idx = result_df.columns.get_loc(target_color_col) + 1
                            for row_num in range(2, len(result_df) + 2):
                                cell = worksheet.cell(row=row_num, column=col_idx)
                                matches = re.findall(r'(\d{1,2}):(\d{2})', str(cell.value))
                                for h, m in matches:
                                    t = time(int(h), int(m))
                                    if start_t and end_t and start_t <= t <= end_t:
                                        cell.fill = yellow_fill

                    st.success("✅ تم الانتهاء من دمج كافة البيانات وتلوين عمود الوقت وإخفاء الأعمدة المطلوبة بنجاح!")
                    st.download_button("📥 تحميل الملف المدمج الصافي", output.getvalue(), "Combined_Report_Cleaned.xlsx")
            
            except Exception as e:
                st.error(f"حدث خطأ أثناء معالجة البيانات: {e}")
        else:
            st.warning("الرجاء رفع الملفين أولاً")

if __name__ == "__main__":
    main()
